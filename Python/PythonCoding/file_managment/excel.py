from zipapp import shebang_encoding

from openpyxl import workbook, load_workbook
from openpyxl.workbook import Workbook

from file_managment.jsonfilemgmnt import filename


def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active

    sheet.append(["Name", "Age"])
    sheet.append(["sindhu", 22])
    sheet.append(["telukuntla", 21])

    workbook.save(filename)
    print('Exceeeeeel file written successfully')

def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        print(f'Name: {row[0]}, Age: {row[1]}')

filename = "data.xlsx"

write_excel(filename)

print("Data read from Excel file :")
read_excel(filename)